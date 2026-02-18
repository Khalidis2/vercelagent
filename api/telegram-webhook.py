# api/telegram-webhook.py
# ─────────────────────────────────────────────────────────────────
# Farm Accounting Bot — Final Build v3
#
# Sheet columns: A=date, B=type, C=item, D=amount, E=user,
#                F=category, G=notes
#
# New features:
#  • Live balance (رصيد لحظي)
#  • Date search (بحث بالتاريخ)
#  • Category — fixed: مواشي, required on every transaction
#  • Notes on transaction (ملاحظات)
#  • Text search across all transactions (بحث نصي)
#
# All previous fixes retained:
#  • Vercel timeout → 200 immediately, background thread
#  • Export timeout → background thread
#  • Empty/command messages → ignored gracefully
#  • sheetId → dynamic lookup by name
#  • Alert spam → once per calendar month
#  • Fuzzy item search → case-insensitive partial match
#  • SHEET_NAME → env var
# ─────────────────────────────────────────────────────────────────

from http.server import BaseHTTPRequestHandler
import json, os, io, logging, threading
from datetime import datetime, timezone, timedelta

import requests
from openai import OpenAI
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER

try:
    import arabic_reshaper
    from bidi.algorithm import get_display
    ARABIC_SUPPORT = True
except ImportError:
    ARABIC_SUPPORT = False

# ─────────────────────────────────────────────────────────────────
# Config
# ─────────────────────────────────────────────────────────────────

logging.basicConfig(level=logging.INFO)
log = logging.getLogger(__name__)

TELEGRAM_BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN")
OPENAI_API_KEY     = os.environ.get("OPENAI_API_KEY")
GOOGLE_SA_JSON     = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
SPREADSHEET_ID     = os.environ.get("SPREADSHEET_ID")
SHEET_NAME         = os.environ.get("SHEET_NAME", "Transactions")

USERS = {
    47329648:   {"name": "Khaled", "role": "admin"},
    6894180427: {"name": "Hamad",  "role": "admin"},
}

MONTHLY_EXPENSE_ALERT = float(os.environ.get("EXPENSE_ALERT", "10000"))

# Fixed categories — "مواشي" is the only one, required on every transaction
CATEGORIES = ["مواشي"]

UAE_TZ  = timezone(timedelta(hours=4))
DIVIDER = "────────────"

_last_alert_month: str = ""

openai_client = OpenAI(api_key=OPENAI_API_KEY)


# ─────────────────────────────────────────────────────────────────
# Telegram
# ─────────────────────────────────────────────────────────────────

def send(chat_id, text):
    try:
        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage",
            json={"chat_id": chat_id, "text": text},
            timeout=15,
        )
    except Exception as e:
        log.error(f"send error: {e}")


def send_document(chat_id, file_bytes, filename, caption=""):
    try:
        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument",
            data={"chat_id": chat_id, "caption": caption},
            files={"document": (filename, file_bytes)},
            timeout=30,
        )
    except Exception as e:
        log.error(f"send_document error: {e}")


def broadcast_admins(text):
    for uid, info in USERS.items():
        if info["role"] == "admin":
            send(uid, text)


# ─────────────────────────────────────────────────────────────────
# Google Sheets
# ─────────────────────────────────────────────────────────────────

def get_service():
    creds = Credentials.from_service_account_info(
        json.loads(GOOGLE_SA_JSON),
        scopes=["https://www.googleapis.com/auth/spreadsheets"],
    )
    return build("sheets", "v4", credentials=creds)


def get_sheet_id(service):
    try:
        meta = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()
        for s in meta.get("sheets", []):
            if s["properties"]["title"] == SHEET_NAME:
                return s["properties"]["sheetId"]
    except Exception as e:
        log.error(f"get_sheet_id: {e}")
    return 0


def load_transactions(service):
    """
    Loads A2:G — columns: date, type, item, amount, user, category, notes
    row_index = real 1-based sheet row (header is row 1)
    """
    try:
        res = (
            service.spreadsheets().values()
            .get(spreadsheetId=SPREADSHEET_ID, range=f"{SHEET_NAME}!A2:G")
            .execute()
        )
        rows = res.get("values", [])
    except Exception as e:
        log.error(f"load_transactions: {e}")
        return []

    data = []
    for i, r in enumerate(rows, start=2):
        if len(r) < 4:
            continue
        data.append({
            "row_index": i,
            "date":     r[0],
            "type":     r[1],
            "item":     r[2],
            "amount":   r[3],
            "user":     r[4] if len(r) > 4 else "",
            "category": r[5] if len(r) > 5 else "",
            "notes":    r[6] if len(r) > 6 else "",
        })
    return data


def append_transaction(service, kind, item, amount, user, category, notes):
    try:
        ts = datetime.now(UAE_TZ).strftime("%Y-%m-%d %H:%M")
        service.spreadsheets().values().append(
            spreadsheetId=SPREADSHEET_ID,
            range=f"{SHEET_NAME}!A1:G1",
            valueInputOption="USER_ENTERED",
            body={"values": [[ts, kind, item, amount, user, category, notes]]},
        ).execute()
        log.info(f"Saved: {kind} | {item} | {amount} | {category} | {user}")
        return True
    except Exception as e:
        log.error(f"append_transaction: {e}")
        return False


def delete_row(service, row_index):
    try:
        sheet_id = get_sheet_id(service)
        body = {
            "requests": [{
                "deleteDimension": {
                    "range": {
                        "sheetId":    sheet_id,
                        "dimension":  "ROWS",
                        "startIndex": row_index - 1,
                        "endIndex":   row_index,
                    }
                }
            }]
        }
        service.spreadsheets().batchUpdate(
            spreadsheetId=SPREADSHEET_ID, body=body
        ).execute()
        return True
    except Exception as e:
        log.error(f"delete_row: {e}")
        return False


def update_row(service, row_index, kind, item, amount, user, category, notes):
    try:
        ts = datetime.now(UAE_TZ).strftime("%Y-%m-%d %H:%M")
        service.spreadsheets().values().update(
            spreadsheetId=SPREADSHEET_ID,
            range=f"{SHEET_NAME}!A{row_index}:G{row_index}",
            valueInputOption="USER_ENTERED",
            body={"values": [[ts, kind, item, amount, user, category, notes]]},
        ).execute()
        return True
    except Exception as e:
        log.error(f"update_row: {e}")
        return False


# ─────────────────────────────────────────────────────────────────
# Aggregation helpers
# ─────────────────────────────────────────────────────────────────

def now_uae():
    return datetime.now(UAE_TZ)


def parse_amount(val):
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0.0


def _parse_date(date_str):
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(date_str[:16], fmt).date()
        except ValueError:
            continue
    return datetime.min.date()


def filter_by_period(transactions, period):
    today = now_uae().date()
    if period == "all":
        return transactions
    if period == "today":
        key = today.isoformat()[:10]
        return [t for t in transactions if t["date"][:10] == key]
    if period == "this_week":
        week_start = today - timedelta(days=today.weekday())
        return [t for t in transactions if _parse_date(t["date"]) >= week_start]
    if period == "this_month":
        prefix = today.strftime("%Y-%m")
        return [t for t in transactions if t["date"].startswith(prefix)]
    if period == "last_month":
        last   = today.replace(day=1) - timedelta(days=1)
        prefix = last.strftime("%Y-%m")
        return [t for t in transactions if t["date"].startswith(prefix)]
    return transactions


def compute_totals(rows):
    income  = sum(parse_amount(r["amount"]) for r in rows if r["type"] == "دخل")
    expense = sum(parse_amount(r["amount"]) for r in rows if r["type"] == "صرف")
    return {"income": income, "expense": expense, "net": income - expense}


def fmt_amount(val):
    f = float(val)
    return f"{int(f):,}" if f.is_integer() else f"{f:,.2f}"


# ─────────────────────────────────────────────────────────────────
# Report builders
# ─────────────────────────────────────────────────────────────────

PERIOD_LABELS = {
    "today":      "اليوم",
    "this_week":  "هذا الأسبوع",
    "this_month": "هذا الشهر",
    "last_month": "الشهر الماضي",
    "all":        "الكل",
}

PERIOD_LABELS_EN = {
    "today":      "Today",
    "this_week":  "This Week",
    "this_month": "This Month",
    "last_month": "Last Month",
    "all":        "All Time",
}


def build_report(transactions, period, show="all"):
    rows  = filter_by_period(transactions, period)
    tots  = compute_totals(rows)
    sign  = "+" if tots["net"] >= 0 else ""
    label = PERIOD_LABELS.get(period, period)

    if show == "income":
        return (
            f"{DIVIDER}\n"
            f"إجمالي الدخل — {label}\n"
            f"{DIVIDER}\n"
            f"الدخل: {fmt_amount(tots['income'])} د.إ\n"
            f"{DIVIDER}"
        )
    if show == "expense":
        return (
            f"{DIVIDER}\n"
            f"إجمالي المصروف — {label}\n"
            f"{DIVIDER}\n"
            f"المصروف: {fmt_amount(tots['expense'])} د.إ\n"
            f"{DIVIDER}"
        )
    if show == "net":
        status = "✅ ربح" if tots["net"] >= 0 else "🔴 خسارة"
        return (
            f"{DIVIDER}\n"
            f"الصافي — {label}\n"
            f"{DIVIDER}\n"
            f"الدخل:    {fmt_amount(tots['income'])} د.إ\n"
            f"المصروف:  {fmt_amount(tots['expense'])} د.إ\n"
            f"الصافي:   {sign}{fmt_amount(tots['net'])} د.إ  {status}\n"
            f"{DIVIDER}"
        )
    return (
        f"{DIVIDER}\n"
        f"تقرير {label}\n"
        f"{DIVIDER}\n"
        f"الدخل:     {fmt_amount(tots['income'])} د.إ\n"
        f"المصروف:   {fmt_amount(tots['expense'])} د.إ\n"
        f"الصافي:    {sign}{fmt_amount(tots['net'])} د.إ\n"
        f"{DIVIDER}"
    )


def build_details(transactions, period, tx_filter="all", limit=10):
    rows  = filter_by_period(transactions, period)
    label = PERIOD_LABELS.get(period, period)

    if tx_filter == "دخل":
        rows = [r for r in rows if r["type"] == "دخل"]
    elif tx_filter == "صرف":
        rows = [r for r in rows if r["type"] == "صرف"]

    rows = list(reversed(rows))[:limit]

    if not rows:
        return f"{DIVIDER}\nلا توجد عمليات مسجلة في هذه الفترة.\n{DIVIDER}"

    lines = [DIVIDER, f"آخر {len(rows)} عملية — {label}", DIVIDER]
    for i, r in enumerate(rows, 1):
        t_label = "✅ دخل" if r["type"] == "دخل" else "🔴 صرف"
        line = f"{i}. {r['date'][:10]} | {t_label} | {r['item']} | {fmt_amount(r['amount'])} د.إ"
        if r.get("category"):
            line += f" | {r['category']}"
        if r.get("notes"):
            line += f"\n   📝 {r['notes']}"
        lines.append(line)
    lines.append(DIVIDER)
    return "\n".join(lines)


def build_comparison(transactions, pa, pb):
    t_a  = compute_totals(filter_by_period(transactions, pa))
    t_b  = compute_totals(filter_by_period(transactions, pb))
    la   = PERIOD_LABELS.get(pa, pa)
    lb   = PERIOD_LABELS.get(pb, pb)
    diff = t_a["net"] - t_b["net"]
    sign = "+" if diff >= 0 else ""

    def block(label, t):
        s = "+" if t["net"] >= 0 else ""
        return (
            f"الفترة: {label}\n"
            f"  الدخل:    {fmt_amount(t['income'])} د.إ\n"
            f"  المصروف:  {fmt_amount(t['expense'])} د.إ\n"
            f"  الصافي:   {s}{fmt_amount(t['net'])} د.إ"
        )

    return (
        f"{DIVIDER}\n"
        f"مقارنة\n"
        f"{DIVIDER}\n"
        f"{block(la, t_a)}\n"
        f"{DIVIDER}\n"
        f"{block(lb, t_b)}\n"
        f"{DIVIDER}\n"
        f"فرق الصافي: {sign}{fmt_amount(diff)} د.إ\n"
        f"{DIVIDER}"
    )


# ─────────────────────────────────────────────────────────────────
# NEW: Live Balance (رصيد لحظي)
# ─────────────────────────────────────────────────────────────────

def build_balance(transactions):
    """إجمالي الدخل والمصروف والصافي من أول عملية حتى الآن."""
    tots   = compute_totals(transactions)
    sign   = "+" if tots["net"] >= 0 else ""
    status = "✅ ربح" if tots["net"] >= 0 else "🔴 خسارة"
    now    = datetime.now(UAE_TZ).strftime("%Y-%m-%d %H:%M")
    count  = len(transactions)

    return (
        f"{DIVIDER}\n"
        f"الرصيد الحالي\n"
        f"{DIVIDER}\n"
        f"الدخل الكلي:    {fmt_amount(tots['income'])} د.إ\n"
        f"المصروف الكلي:  {fmt_amount(tots['expense'])} د.إ\n"
        f"الصافي:         {sign}{fmt_amount(tots['net'])} د.إ  {status}\n"
        f"{DIVIDER}\n"
        f"عدد العمليات:   {count}\n"
        f"آخر تحديث:      {now}\n"
        f"{DIVIDER}"
    )


# ─────────────────────────────────────────────────────────────────
# NEW: Date Search (بحث بالتاريخ)
# ─────────────────────────────────────────────────────────────────

def build_date_search(transactions, date_str):
    """
    يبحث عن عمليات يوم معين.
    date_str: YYYY-MM-DD أو اليوم أو أمس
    """
    today = now_uae().date()

    if date_str in ("اليوم", "today"):
        target = today
    elif date_str in ("أمس", "yesterday"):
        target = today - timedelta(days=1)
    else:
        try:
            target = datetime.strptime(date_str[:10], "%Y-%m-%d").date()
        except ValueError:
            return f"{DIVIDER}\nصيغة التاريخ غير صحيحة. مثال: 2024-03-15\n{DIVIDER}"

    key  = target.isoformat()
    rows = [t for t in transactions if t["date"][:10] == key]

    if not rows:
        return f"{DIVIDER}\nلا توجد عمليات في {key}\n{DIVIDER}"

    tots  = compute_totals(rows)
    sign  = "+" if tots["net"] >= 0 else ""
    lines = [DIVIDER, f"عمليات يوم {key}", DIVIDER]

    for i, r in enumerate(rows, 1):
        t_label = "✅ دخل" if r["type"] == "دخل" else "🔴 صرف"
        line = f"{i}. {r['date'][11:16]} | {t_label} | {r['item']} | {fmt_amount(r['amount'])} د.إ"
        if r.get("category"):
            line += f" | {r['category']}"
        if r.get("notes"):
            line += f"\n   📝 {r['notes']}"
        lines.append(line)

    lines += [
        DIVIDER,
        f"الدخل:    {fmt_amount(tots['income'])} د.إ",
        f"المصروف:  {fmt_amount(tots['expense'])} د.إ",
        f"الصافي:   {sign}{fmt_amount(tots['net'])} د.إ",
        DIVIDER,
    ]
    return "\n".join(lines)


# ─────────────────────────────────────────────────────────────────
# NEW: Text Search (بحث نصي)
# ─────────────────────────────────────────────────────────────────

def build_text_search(transactions, query, limit=20):
    """
    يبحث في اسم البند والملاحظات والتصنيف.
    case-insensitive partial match.
    """
    needle = query.strip().lower()
    if not needle:
        return f"{DIVIDER}\nيرجى إدخال كلمة للبحث.\n{DIVIDER}"

    matches = [
        t for t in transactions
        if (needle in t["item"].lower()
            or needle in t.get("notes", "").lower()
            or needle in t.get("category", "").lower())
    ]

    if not matches:
        return f"{DIVIDER}\nلا توجد نتائج لـ: {query}\n{DIVIDER}"

    rows  = list(reversed(matches))[:limit]
    tots  = compute_totals(matches)
    sign  = "+" if tots["net"] >= 0 else ""

    lines = [DIVIDER, f"نتائج البحث: {query} ({len(matches)} عملية)", DIVIDER]
    for i, r in enumerate(rows, 1):
        t_label = "✅ دخل" if r["type"] == "دخل" else "🔴 صرف"
        line = f"{i}. {r['date'][:10]} | {t_label} | {r['item']} | {fmt_amount(r['amount'])} د.إ"
        if r.get("category"):
            line += f" | {r['category']}"
        if r.get("notes"):
            line += f"\n   📝 {r['notes']}"
        lines.append(line)

    if len(matches) > limit:
        lines.append(f"... و{len(matches) - limit} عملية أخرى")

    lines += [
        DIVIDER,
        f"إجمالي الدخل:    {fmt_amount(tots['income'])} د.إ",
        f"إجمالي المصروف:  {fmt_amount(tots['expense'])} د.إ",
        f"الصافي:          {sign}{fmt_amount(tots['net'])} د.إ",
        DIVIDER,
    ]
    return "\n".join(lines)


# ─────────────────────────────────────────────────────────────────
# Edit / Delete
# ─────────────────────────────────────────────────────────────────

def _find_row(transactions, target, item_name):
    if target == "last":
        return transactions[-1] if transactions else None
    if not item_name:
        return None
    needle  = item_name.strip().lower()
    matches = [t for t in transactions if needle in t["item"].strip().lower()]
    return matches[-1] if matches else None


def handle_edit(service, transactions, intent_data, user_name):
    target    = intent_data.get("target", "last")
    item_name = intent_data.get("item_name") or ""
    action    = intent_data.get("action", "delete")

    row = _find_row(transactions, target, item_name)
    if not row:
        if target == "last":
            return "لا توجد عمليات مسجلة."
        return f"لم يتم العثور على عملية تحتوي على: {item_name}"

    row_index = row["row_index"]

    if action == "delete":
        ok = delete_row(service, row_index)
        if ok:
            return (
                f"{DIVIDER}\n"
                f"تم الحذف\n"
                f"{DIVIDER}\n"
                f"البند:    {row['item']}\n"
                f"المبلغ:   {fmt_amount(row['amount'])} د.إ\n"
                f"التاريخ:  {row['date'][:10]}\n"
                f"{DIVIDER}"
            )
        return "⚠️ حدث خطأ أثناء الحذف."

    # update
    new_amount   = intent_data.get("new_amount")
    new_item     = intent_data.get("new_item")     or row["item"]
    new_type     = intent_data.get("new_type")     or row["type"]
    new_category = intent_data.get("new_category") or row.get("category", "")
    new_notes    = intent_data.get("new_notes")    or row.get("notes", "")

    try:
        new_amount = abs(float(new_amount)) if new_amount else parse_amount(row["amount"])
    except (ValueError, TypeError):
        return "⚠️ المبلغ الجديد غير صالح."

    ok = update_row(service, row_index, new_type, new_item,
                    new_amount, user_name, new_category, new_notes)
    if ok:
        return (
            f"{DIVIDER}\n"
            f"تم التعديل\n"
            f"{DIVIDER}\n"
            f"البند:          {new_item}\n"
            f"المبلغ الجديد:  {fmt_amount(new_amount)} د.إ\n"
            f"{DIVIDER}"
        )
    return "⚠️ حدث خطأ أثناء التعديل."


# ─────────────────────────────────────────────────────────────────
# Alert
# ─────────────────────────────────────────────────────────────────

def check_expense_alert(transactions):
    global _last_alert_month
    monthly       = filter_by_period(transactions, "this_month")
    expense       = compute_totals(monthly)["expense"]
    current_month = now_uae().strftime("%Y-%m")

    if expense >= MONTHLY_EXPENSE_ALERT and current_month != _last_alert_month:
        _last_alert_month = current_month
        broadcast_admins(
            f"⚠️ تنبيه: المصروف الشهري تجاوز الحد\n"
            f"{DIVIDER}\n"
            f"الإجمالي هذا الشهر: {fmt_amount(expense)} د.إ\n"
            f"الحد المحدد:        {fmt_amount(MONTHLY_EXPENSE_ALERT)} د.إ\n"
            f"{DIVIDER}"
        )


# ─────────────────────────────────────────────────────────────────
# Arabic text helper for PDF
# ─────────────────────────────────────────────────────────────────

def ar(text):
    if not ARABIC_SUPPORT:
        return text
    try:
        reshaped = arabic_reshaper.reshape(str(text))
        return get_display(reshaped)
    except Exception:
        return text


# ─────────────────────────────────────────────────────────────────
# Excel Export — updated for new columns (F=category, G=notes)
# ─────────────────────────────────────────────────────────────────

def build_excel(transactions, period):
    rows  = filter_by_period(transactions, period)
    tots  = compute_totals(rows)
    label = PERIOD_LABELS.get(period, period)

    wb     = openpyxl.Workbook()
    thin   = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb.active
    ws.title = "Transactions"
    ws.sheet_view.rightToLeft = True

    h_fill = PatternFill("solid", fgColor="1F4E79")
    i_fill = PatternFill("solid", fgColor="E2EFDA")
    e_fill = PatternFill("solid", fgColor="FCE4D6")

    headers    = ["التاريخ", "النوع", "البند", "المبلغ (د.إ)", "المستخدم", "التصنيف", "ملاحظات"]
    col_widths = [18, 10, 28, 16, 14, 14, 30]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = Font(bold=True, color="FFFFFF", size=11)
        c.fill      = h_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[1].height = 22

    for i, r in enumerate(rows, 2):
        fill = i_fill if r["type"] == "دخل" else e_fill
        vals = [r["date"], r["type"], r["item"], parse_amount(r["amount"]),
                r["user"], r.get("category", ""), r.get("notes", "")]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.fill      = fill
            c.alignment = Alignment(horizontal="right", wrap_text=(j == 7))
            c.border    = border
            if j == 4:
                c.number_format = "#,##0.00"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.rightToLeft = True
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 18
    s_fill  = PatternFill("solid", fgColor="D6E4F0")
    summary = [
        ("الفترة",       label),
        ("الدخل",        tots["income"]),
        ("المصروف",      tots["expense"]),
        ("الصافي",       tots["net"]),
        ("عدد العمليات", len(rows)),
    ]
    for i, (k, v) in enumerate(summary, 1):
        ck = ws2.cell(row=i, column=1, value=k)
        cv = ws2.cell(row=i, column=2, value=v)
        for c in (ck, cv):
            c.fill      = s_fill
            c.border    = border
            c.alignment = Alignment(horizontal="right")
        ck.font = Font(bold=True, size=11)
        if isinstance(v, float):
            cv.number_format = "#,##0.00"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────
# PDF helpers
# ─────────────────────────────────────────────────────────────────

def _table_style_header():
    return TableStyle([
        ("BACKGROUND",     (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR",      (0, 0), (-1, 0), colors.white),
        ("FONTNAME",       (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE",       (0, 0), (-1, 0), 11),
        ("FONTSIZE",       (0, 1), (-1, -1), 10),
        ("ALIGN",          (0, 0), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.HexColor("#EBF5FB"), colors.white]),
        ("GRID",           (0, 0), (-1, -1), 0.5, colors.HexColor("#AAAAAA")),
        ("BOTTOMPADDING",  (0, 0), (-1, -1), 6),
        ("TOPPADDING",     (0, 0), (-1, -1), 6),
    ])


def _table_style_detail():
    return TableStyle([
        ("BACKGROUND",     (0, 0), (-1, 0), colors.HexColor("#2E86C1")),
        ("TEXTCOLOR",      (0, 0), (-1, 0), colors.white),
        ("FONTNAME",       (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE",       (0, 0), (-1, 0), 9),
        ("FONTSIZE",       (0, 1), (-1, -1), 8),
        ("ALIGN",          (0, 0), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.HexColor("#EBF5FB"), colors.white]),
        ("GRID",           (0, 0), (-1, -1), 0.4, colors.HexColor("#AAAAAA")),
        ("BOTTOMPADDING",  (0, 0), (-1, -1), 4),
        ("TOPPADDING",     (0, 0), (-1, -1), 4),
    ])


# ─────────────────────────────────────────────────────────────────
# PDF — Arabic (updated columns)
# ─────────────────────────────────────────────────────────────────

def build_pdf_arabic(transactions, period):
    rows   = filter_by_period(transactions, period)
    tots   = compute_totals(rows)
    sign   = "+" if tots["net"] >= 0 else ""
    label  = PERIOD_LABELS.get(period, period)
    issued = datetime.now(UAE_TZ).strftime("%Y-%m-%d %H:%M")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles  = getSampleStyleSheet()
    title_s = ParagraphStyle("t",  parent=styles["Title"],    fontSize=16, spaceAfter=4,  alignment=TA_CENTER)
    sub_s   = ParagraphStyle("s",  parent=styles["Normal"],   fontSize=9,  spaceAfter=10, alignment=TA_CENTER, textColor=colors.grey)
    h2_s    = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=11, spaceBefore=8)

    story = [
        Paragraph(ar(f"تقرير المزرعة — {label}"), title_s),
        Paragraph(ar(f"تاريخ الإصدار: {issued}"), sub_s),
        Spacer(1, 0.3*cm),
    ]

    summary_data = [
        [ar("البيان"),   ar("المبلغ (د.إ)")],
        [ar("الدخل"),    fmt_amount(tots["income"])],
        [ar("المصروف"),  fmt_amount(tots["expense"])],
        [ar("الصافي"),   f"{sign}{fmt_amount(tots['net'])}"],
    ]
    st = Table(summary_data, colWidths=[8*cm, 6*cm])
    st.setStyle(_table_style_header())
    story += [st, Spacer(1, 0.5*cm)]

    if rows:
        story.append(Paragraph(ar("تفاصيل العمليات"), h2_s))
        td = [[ar("التاريخ"), ar("النوع"), ar("البند"),
               ar("المبلغ (د.إ)"), ar("التصنيف"), ar("ملاحظات")]]
        for r in rows:
            td.append([
                r["date"][:10],
                ar(r["type"]),
                ar(r["item"]),
                fmt_amount(r["amount"]),
                ar(r.get("category", "")),
                ar(r.get("notes", "")),
            ])
        t = Table(td, colWidths=[2.8*cm, 1.8*cm, 5.5*cm, 2.8*cm, 2*cm, 3*cm], repeatRows=1)
        t.setStyle(_table_style_detail())
        story.append(t)

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────
# PDF — English (updated columns)
# ─────────────────────────────────────────────────────────────────

def build_pdf_english(transactions, period):
    rows   = filter_by_period(transactions, period)
    tots   = compute_totals(rows)
    sign   = "+" if tots["net"] >= 0 else ""
    label  = PERIOD_LABELS_EN.get(period, period)
    issued = datetime.now(UAE_TZ).strftime("%Y-%m-%d %H:%M")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles  = getSampleStyleSheet()
    title_s = ParagraphStyle("t",  parent=styles["Title"],    fontSize=16, spaceAfter=4,  alignment=TA_CENTER)
    sub_s   = ParagraphStyle("s",  parent=styles["Normal"],   fontSize=9,  spaceAfter=10, alignment=TA_CENTER, textColor=colors.grey)
    h2_s    = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=11, spaceBefore=8)

    story = [
        Paragraph(f"Farm Report — {label}", title_s),
        Paragraph(f"Issued: {issued} (UAE)", sub_s),
        Spacer(1, 0.3*cm),
    ]

    summary_data = [
        ["Item",    "Amount (AED)"],
        ["Income",  fmt_amount(tots["income"])],
        ["Expense", fmt_amount(tots["expense"])],
        ["Net",     f"{sign}{fmt_amount(tots['net'])}"],
    ]
    st = Table(summary_data, colWidths=[8*cm, 6*cm])
    st.setStyle(_table_style_header())
    story += [st, Spacer(1, 0.5*cm)]

    if rows:
        story.append(Paragraph("Transaction Details", h2_s))
        td = [["Date", "Type", "Item", "Amount (AED)", "Category", "Notes"]]
        for r in rows:
            type_en = "Income" if r["type"] == "دخل" else "Expense"
            td.append([
                r["date"][:10], type_en, r["item"],
                fmt_amount(r["amount"]),
                r.get("category", ""),
                r.get("notes", ""),
            ])
        t = Table(td, colWidths=[2.8*cm, 2*cm, 5*cm, 2.8*cm, 2*cm, 3*cm], repeatRows=1)
        t.setStyle(_table_style_detail())
        story.append(t)

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────
# Export handler — background thread
# ─────────────────────────────────────────────────────────────────

def _export_worker(chat_id, transactions, intent_data):
    period   = intent_data.get("period", "this_month")
    fmt      = intent_data.get("fmt", "both")
    pdf_lang = intent_data.get("pdf_lang", "both")
    ts_str   = datetime.now(UAE_TZ).strftime("%Y-%m-%d")

    if fmt in ("excel", "both"):
        try:
            xlsx = build_excel(transactions, period)
            send_document(chat_id, xlsx,
                          f"Farm_Report_{ts_str}.xlsx",
                          caption=f"📊 Excel — {PERIOD_LABELS.get(period, period)}")
        except Exception as e:
            log.error(f"Excel export: {e}")
            send(chat_id, "⚠️ خطأ في إنشاء ملف Excel.")

    if fmt in ("pdf", "both"):
        if pdf_lang in ("ar", "both"):
            try:
                pdf = build_pdf_arabic(transactions, period)
                send_document(chat_id, pdf,
                              f"Farm_Report_AR_{ts_str}.pdf",
                              caption=f"📄 PDF عربي — {PERIOD_LABELS.get(period, period)}")
            except Exception as e:
                log.error(f"PDF Arabic: {e}")
                send(chat_id, "⚠️ خطأ في إنشاء PDF العربي.")

        if pdf_lang in ("en", "both"):
            try:
                pdf = build_pdf_english(transactions, period)
                send_document(chat_id, pdf,
                              f"Farm_Report_EN_{ts_str}.pdf",
                              caption=f"📄 PDF English — {PERIOD_LABELS_EN.get(period, period)}")
            except Exception as e:
                log.error(f"PDF English: {e}")
                send(chat_id, "⚠️ خطأ في إنشاء PDF الإنجليزي.")


def handle_export(chat_id, transactions, intent_data):
    send(chat_id, "⏳ جاري إنشاء الملف...")
    threading.Thread(
        target=_export_worker,
        args=(chat_id, transactions, intent_data),
        daemon=True,
    ).start()


# ─────────────────────────────────────────────────────────────────
# AI Engine
# ─────────────────────────────────────────────────────────────────

SYSTEM_PROMPT = """
أنت محلل نية (intent classifier) لبوت محاسبة عزبة في الإمارات.

مهمتك الوحيدة: تحليل رسالة المستخدم وإعادة JSON يصف نيته بدقة.
لا تحسب أرقاماً. لا تنشئ تقارير. الكود سيتولى ذلك.

قواعد صارمة:
- أعد JSON فقط. لا نص خارجه. لا Markdown. لا ```.
- لا تخترع أرقاماً أبداً.
- اللغة العربية بكل لهجاتها مدعومة.
- التصنيف الوحيد المتاح: مواشي (إلزامي في كل عملية).

══════════════════════════════════════════
الأنواع الممكنة:
══════════════════════════════════════════

1. عملية مالية — category إلزامي دائماً = "مواشي":
{"intent":"transaction","type":"دخل|صرف","item":"اسم البند","amount":<رقم>,"date":"اليوم|أمس|<تاريخ>","category":"مواشي","notes":"<ملاحظة أو null>"}

2. تقرير:
{"intent":"report","period":"today|this_week|this_month|last_month|all","show":"income|expense|net|all"}

3. تفاصيل:
{"intent":"details","period":"today|this_week|this_month|last_month|all","filter":"all|دخل|صرف","limit":<عدد أو null>}

4. مقارنة:
{"intent":"comparison","period_a":"this_week|this_month|last_month|all","period_b":"this_week|this_month|last_month|all"}

5. ملخص أسبوعي:
{"intent":"weekly_summary"}

6. ملخص شهري:
{"intent":"monthly_summary"}

7. رصيد لحظي:
{"intent":"balance"}

8. بحث بالتاريخ:
{"intent":"date_search","date":"YYYY-MM-DD|اليوم|أمس"}

9. بحث نصي:
{"intent":"text_search","query":"<كلمة البحث>","limit":<عدد أو null>}

10. تصدير:
{"intent":"export","period":"today|this_week|this_month|last_month|all","fmt":"excel|pdf|both","pdf_lang":"ar|en|both"}

11. حذف:
{"intent":"edit","action":"delete","target":"last|item","item_name":"<اسم أو null>"}

12. تعديل:
{"intent":"edit","action":"update","target":"last|item","item_name":"<اسم أو null>","new_amount":<رقم أو null>,"new_item":"<اسم جديد أو null>","new_type":"دخل|صرف|null","new_notes":"<ملاحظة جديدة أو null>"}

13. محادثة:
{"intent":"conversation","reply":"<رد مختصر رسمي>"}

══════════════════════════════════════════
أمثلة:
══════════════════════════════════════════

"بعنا قمح بـ 3000"                      → transaction / دخل / مواشي / notes=null
"بعنا قمح 3000 ملاحظة دفعة أولى"       → transaction / دخل / مواشي / notes=دفعة أولى
"دفعنا كهرباء 500"                      → transaction / صرف / مواشي / notes=null
"كم صرفنا هالشهر؟"                     → report / this_month / show=expense
"كم جبنا اليوم؟"                        → report / today / show=income
"هل نحن في ربح؟"                        → report / all / show=net
"عطني تقرير كامل"                       → report / this_month / show=all
"كم رصيدنا الحين؟"                      → balance
"وين وصلنا؟"                            → balance
"عمليات يوم 2024-03-15"                 → date_search / 2024-03-15
"عمليات أمس"                            → date_search / أمس
"كل عمليات القمح"                       → text_search / query=قمح
"دوّر كهرباء"                           → text_search / query=كهرباء
"آخر 5 عمليات"                          → details / all / limit=5
"ملخص الأسبوع"                          → weekly_summary
"قارن هالشهر بالشهر الماضي"             → comparison
"صدّر تقرير هالشهر"                     → export / this_month / fmt=both / pdf_lang=both
"ابعثلي Excel"                          → export / this_month / fmt=excel
"PDF عربي"                              → export / this_month / fmt=pdf / pdf_lang=ar
"احذف آخر عملية"                        → edit / delete / last
"احذف الكهرباء"                         → edit / delete / item / كهرباء
"عدّل الكهرباء المبلغ 300"              → edit / update / item / كهرباء / new_amount=300
"صباح الخير"                            → conversation
""".strip()


def ask_ai(user_text):
    try:
        resp = openai_client.chat.completions.create(
            model="gpt-4o-mini",
            temperature=0,
            max_tokens=350,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user",   "content": user_text},
            ],
        )
        raw = resp.choices[0].message.content or ""
        log.info(f"AI: {raw[:200]}")
        return _parse_ai(raw)
    except Exception as e:
        log.error(f"OpenAI error: {e}")
        return _fallback("حدث خطأ في الاتصال بالذكاء الاصطناعي.")


def _parse_ai(raw):
    text = raw.strip()
    if "```" in text:
        text = "\n".join(
            l for l in text.splitlines() if not l.strip().startswith("```")
        ).strip()
    data = None
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        s, e = text.find("{"), text.rfind("}") + 1
        if s != -1 and e > s:
            try:
                data = json.loads(text[s:e])
            except json.JSONDecodeError:
                pass
    if data is None:
        log.warning(f"JSON parse failed: {text[:150]}")
        return _fallback("لم أستطع فهم الرسالة. يرجى إعادة الصياغة.")
    return _validate_ai(data)


def _validate_ai(data):
    intent = data.get("intent")

    if intent == "transaction":
        if not data.get("type") or not data.get("item"):
            return _fallback("بيانات العملية غير مكتملة.")
        try:
            data["amount"] = abs(float(data["amount"]))
        except (ValueError, TypeError):
            return _fallback("المبلغ غير صالح. يرجى إدخال رقم.")
        if data["type"] not in ("دخل", "صرف"):
            return _fallback("نوع العملية غير معروف.")
        # enforce category
        data["category"] = "مواشي"
        # normalize notes
        if not data.get("notes") or str(data.get("notes")).lower() in ("null", "none", ""):
            data["notes"] = ""
        return data

    if intent == "edit":
        if data.get("action") not in ("delete", "update"):
            return _fallback("نوع التعديل غير معروف.")
        return data

    if intent == "text_search":
        if not data.get("query"):
            return _fallback("يرجى تحديد كلمة للبحث.")
        return data

    if intent == "date_search":
        if not data.get("date"):
            return _fallback("يرجى تحديد تاريخ.")
        return data

    if intent in ("report", "details", "comparison", "export", "balance",
                  "weekly_summary", "monthly_summary", "conversation"):
        return data

    return _fallback("لم أفهم المطلوب.")


def _fallback(msg):
    return {"intent": "conversation", "reply": msg}


# ─────────────────────────────────────────────────────────────────
# Reply builder
# ─────────────────────────────────────────────────────────────────

def build_reply(intent_data, transactions, user_name, service, chat_id):
    intent = intent_data.get("intent")

    # ── عملية مالية ──────────────────────────────────────────────
    if intent == "transaction":
        kind     = intent_data["type"]
        item     = intent_data["item"]
        amount   = intent_data["amount"]
        date     = intent_data.get("date", "اليوم")
        category = intent_data.get("category", "مواشي")
        notes    = intent_data.get("notes", "")

        if not append_transaction(service, kind, item, amount, user_name, category, notes):
            return "⚠️ حدث خطأ أثناء الحفظ. يرجى المحاولة مرة أخرى."

        if kind == "صرف":
            check_expense_alert(load_transactions(service))

        type_label = "✅ دخل" if kind == "دخل" else "🔴 صرف"
        reply = (
            f"{DIVIDER}\n"
            f"تم التسجيل\n"
            f"{DIVIDER}\n"
            f"التاريخ:    {date}\n"
            f"النوع:      {type_label}\n"
            f"البند:      {item}\n"
            f"المبلغ:     {fmt_amount(amount)} د.إ\n"
            f"التصنيف:    {category}\n"
            f"المستخدم:   {user_name}\n"
        )
        if notes:
            reply += f"ملاحظات:    {notes}\n"
        reply += DIVIDER
        return reply

    # ── تقرير ────────────────────────────────────────────────────
    if intent == "report":
        return build_report(transactions,
                            intent_data.get("period", "all"),
                            intent_data.get("show", "all"))

    # ── تفاصيل ───────────────────────────────────────────────────
    if intent == "details":
        try:
            limit = int(intent_data.get("limit") or 10)
        except (ValueError, TypeError):
            limit = 10
        return build_details(transactions,
                             intent_data.get("period", "all"),
                             intent_data.get("filter", "all"),
                             limit)

    # ── مقارنة ───────────────────────────────────────────────────
    if intent == "comparison":
        return build_comparison(transactions,
                                intent_data.get("period_a", "this_month"),
                                intent_data.get("period_b", "last_month"))

    # ── ملخص ─────────────────────────────────────────────────────
    if intent == "weekly_summary":
        return build_report(transactions, "this_week", "all")
    if intent == "monthly_summary":
        return build_report(transactions, "this_month", "all")

    # ── رصيد لحظي ────────────────────────────────────────────────
    if intent == "balance":
        return build_balance(transactions)

    # ── بحث بالتاريخ ─────────────────────────────────────────────
    if intent == "date_search":
        return build_date_search(transactions, intent_data.get("date", "اليوم"))

    # ── بحث نصي ──────────────────────────────────────────────────
    if intent == "text_search":
        try:
            limit = int(intent_data.get("limit") or 20)
        except (ValueError, TypeError):
            limit = 20
        return build_text_search(transactions, intent_data.get("query", ""), limit)

    # ── تصدير ────────────────────────────────────────────────────
    if intent == "export":
        handle_export(chat_id, transactions, intent_data)
        return None

    # ── تعديل / حذف ──────────────────────────────────────────────
    if intent == "edit":
        return handle_edit(service, transactions, intent_data, user_name)

    # ── محادثة ───────────────────────────────────────────────────
    if intent == "conversation":
        return intent_data.get("reply", "أنا هنا للمساعدة.")

    return "لم أفهم المطلوب."


# ─────────────────────────────────────────────────────────────────
# Main processing — 200 immediately, process in background
# ─────────────────────────────────────────────────────────────────

def _process(update):
    try:
        msg = update.get("message")
        if not msg:
            return

        text = msg.get("text", "").strip()
        if not text or text.startswith("/"):
            return

        chat_id = msg["chat"]["id"]
        user_id = msg["from"]["id"]

        user_info = USERS.get(user_id)
        if not user_info:
            send(chat_id, "غير مصرح.")
            return

        user_name = user_info["name"]
        user_role = user_info["role"]

        try:
            service      = get_service()
            transactions = load_transactions(service)
        except Exception as e:
            log.error(f"Sheets error: {e}")
            send(chat_id, "⚠️ تعذّر الاتصال بقاعدة البيانات.")
            return

        intent_data = ask_ai(text)
        intent      = intent_data.get("intent")

        if intent in ("transaction", "edit") and user_role != "admin":
            send(chat_id, "⛔ ليس لديك صلاحية لهذه العملية.")
            return

        reply = build_reply(intent_data, transactions, user_name, service, chat_id)
        if reply:
            send(chat_id, reply)

    except Exception as e:
        log.error(f"_process error: {e}", exc_info=True)


# ─────────────────────────────────────────────────────────────────
# Webhook Handler
# ─────────────────────────────────────────────────────────────────

class handler(BaseHTTPRequestHandler):

    def _ok(self):
        self.send_response(200)
        self.end_headers()
        self.wfile.write(b"OK")

    def do_GET(self):
        self._ok()

    def do_POST(self):
        try:
            body   = self.rfile.read(int(self.headers.get("Content-Length", 0))).decode()
            update = json.loads(body)
        except Exception as e:
            log.error(f"Bad request: {e}")
            self._ok()
            return

        # Return 200 to Telegram immediately, process in background
        self._ok()
        threading.Thread(target=_process, args=(update,), daemon=True).start()
